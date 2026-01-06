from openpyxl import load_workbook

import argparse
import json
import time
from typing import Dict, List, Tuple, Optional
import re
import sys
from pathlib import Path
# --- Translation constants and helpers (inlined from translate.py) ---

LANGS_DEFAULT = [
    "CS","DE","ES","FR","IT","JA","KO","PL","PT","PTBR","RU","TH","TR","UK","NL","ZH","ZHTW",
]

DOMAIN_FILES = [
    ("ItemName", "ItemName"),
    ("Recipes", "Recipes"),
]

TABLE_HEADER_RE = re.compile(r"^\s*([A-Za-z0-9_]+)\s*=\s*\{\s*$")
ENTRY_RE = re.compile(
    r'''
    ^\s*
    (?:                             # either:
        \[\s*"([^"]+)"\s*\]         #   ["string.key"]
      |                             # or
        ([A-Za-z0-9_.]+)            #   bare_identifier
    )
    \s*=\s*
    "(.*)"
    \s*,?\s*$
    ''',
    re.VERBOSE,
)
TABLE_END_RE = re.compile(r"^\s*\}\s*$")

def read_domain_table(path: Path):
    if not path.exists():
        raise FileNotFoundError(f"Missing translation template: {path}")

    table_name = None
    entries: Dict[str, str] = {}
    in_table = False

    for line in path.read_text(encoding="utf-8").splitlines():
        if not in_table:
            m = TABLE_HEADER_RE.match(line)
            if m:
                table_name = m.group(1)
                in_table = True
            continue

        if TABLE_END_RE.match(line):
            break

        m = ENTRY_RE.match(line)
        if m:
            key = m.group(1) or m.group(2)
            val = m.group(3).replace('\\"', '"')
            entries[key] = val

    if not table_name:
        raise ValueError(f"Could not find table header in {path}")

    domain = table_name.split("_", 1)[0]
    if domain == "Recipe":
        raise ValueError(f"Invalid singular Recipe table in {path}")

    return domain, table_name, entries


def write_domain_table(out_path: Path, domain: str, lang: str, entries: Dict[str, str]):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    table_name = f"{domain}_{lang}"

    lines: List[str] = []
    lines.append(f"{table_name} = {{")
    for key in sorted(entries.keys()):
        val = entries[key].replace('"', '\\"')
        lines.append(f'    ["{key}"] = "{val}",')
    lines.append("}")
    lines.append("")

    out_path.write_text("\n".join(lines), encoding="utf-8")

try:
    from openai import OpenAI
except Exception:
    OpenAI = None

def build_translation_prompt(domain: str, lang: str, entries: Dict[str, str]) -> str:
    context = f"""
You are translating UI strings for a Project Zomboid Build 41 mod.

Rules:
1) DO NOT translate keys.
2) Preserve placeholders like %1 exactly.
3) Output JSON only.
"""
    payload = {
        "domain": domain,
        "target_language": lang,
        "entries": entries,
    }
    return context.strip() + "\n\nDATA:\n" + json.dumps(payload, ensure_ascii=False)

def call_openai_translate(client, model: str, prompt: str) -> Dict[str, str]:
    resp = client.responses.create(
        model=model,
        input=[
            {"role": "developer", "content": "Translate UI strings. Output JSON only."},
            {"role": "user", "content": prompt},
        ],
    )

    text = getattr(resp, "output_text", str(resp)).strip()
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)

    data = json.loads(text)
    if not isinstance(data, dict):
        raise ValueError("Model output was not a JSON object")
    return {str(k): str(v) for k, v in data.items()}

# Excel loader for all sheets
def load_xlsx_rows(path: Path, sheet_name: str) -> list[dict]:
    if not path.exists():
        die(f"Excel file not found: {path}")

    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        die(f"Sheet '{sheet_name}' not found in {path}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    headers = [str(h).strip() for h in rows[0]]
    data = []

    for r in rows[1:]:
        if all(c is None for c in r):
            continue
        row = {}
        for i, h in enumerate(headers):
            val = r[i] if i < len(r) else None
            row[h] = "" if val is None else str(val).strip()
        data.append(row)

    return data

# Containerized loader from Excel
def load_containerized_sheet(xlsx: Path):
    rows = []
    for r in load_xlsx_rows(xlsx, "containerized"):
        item_id = r["item_id"].strip()
        byp = (r.get("byproducts") or "").strip()
        byproducts = [b.strip() for b in byp.split(";") if b.strip()] if byp else []
        rows.append((item_id, byproducts))
    return rows
def emit_containerized_combine_recipe(item_id: str) -> str:
    # item_id like "Base.CannedCornOpen"
    short = item_id.split(".")[-1]
    lines = []
    lines.append(f"    recipe Combine{short}")
    lines.append("    {")
    lines.append(f"        {item_id};2,")
    lines.append(f"        Result : {item_id}=1,")
    lines.append("        Time : 50,")
    lines.append("        Category : Cooking,")
    lines.append("        CanBeDoneFromFloor : true,")
    lines.append("        StopOnWalk        : false,")
    lines.append("        NeedToBeLearn     : false,")
    lines.append("        OnCanPerform : Recipe.OnCanPerform.KitchenConsolidation_Combine_OnCanPerform,")
    lines.append("        OnCreate : Recipe.OnCreate.KitchenConsolidation_Combine_OnCreate,")
    lines.append("    }")
    return "\n".join(lines)
def emit_containerized_lua(rows: list[tuple[str, list[str]]]) -> str:
    lines = []
    lines.append("-- AUTO-GENERATED FILE. DO NOT EDIT.")
    lines.append("RecipeContainerized = RecipeContainerized or {}")
    lines.append("")
    lines.append("local lookup = {")
    for item_id, byps in rows:
        if byps:
            arr = ", ".join([f'\"{b}\"' for b in byps])
            lines.append(f"    [\"{item_id}\"] = {{ {arr} }},")
        else:
            lines.append(f"    [\"{item_id}\"] = {{ }},")
    lines.append("}")
    lines.append("")
    lines.append("function RecipeContainerized.byproductLookup(itemId)")
    lines.append("    return lookup[itemId]")
    lines.append("end")
    return "\n".join(lines)
#!/usr/bin/env python3

import csv
import sys
import re
from pathlib import Path

# NOTE:
# - piece_full_type is derived as "<module>.<piece_name>"
# - display_name_key is derived as "ItemName_<module>.<piece_name>"

USAGE = """
Usage:
  gen-items.py <food-defs.csv> <food-evolved.csv> <mod-root>

Example:
  python gen-items.py food-defs.csv food-evolved-vanilla.csv kitchenconsolidation/media
"""


def die(msg: str):
    print(f"ERROR: {msg}", file=sys.stderr)
    print(USAGE, file=sys.stderr)
    sys.exit(1)


def dash_is_none(value: str):
    if value is None:
        return None
    value = value.strip()
    return None if value == "-" or value == "" else value


def piece_full_type(row: dict) -> str:
    return f"{row['module']}.{row['piece_name']}"


def display_name_key(row: dict) -> str:
    return f"ItemName_{piece_full_type(row)}"


# Helper for evolved item display name key
def evolved_item_display_key(full_type: str, evolved_name: str) -> str:
    # ItemName_KitchenConsolidation.FishPieces_Stew
    return f"ItemName_{full_type}_{evolved_name}"


def humanize_piece_name(piece_name: str) -> str:
    # FishPieces -> Fish Pieces
    out = []
    for i, ch in enumerate(piece_name):
        if i > 0 and ch.isupper() and not piece_name[i - 1].isupper():
            out.append(" ")
        out.append(ch)
    return "".join(out)

# Helper to format evolved item names for pot/pan recipes
def format_pot_pan_evolved_name(base_name: str, evolved_name: str) -> str:
    # evolved_name examples: RicePot, RicePan, PastaPot, PastaPan
    if evolved_name.endswith("Pot"):
        food = evolved_name[:-3]
        return f"Pot of {food} and {base_name}"
    if evolved_name.endswith("Pan"):
        food = evolved_name[:-3]
        return f"Pan of {food} and {base_name}"
    return f"{base_name} {evolved_name}"

# Improved humanize_item_id to normalize underscores
def humanize_item_id(item_id: str) -> str:
    # Base.CannedBellPepper_Open -> Canned Bell Pepper Open
    short = item_id.split(".")[-1].replace("_", " ")
    out = []
    for i, ch in enumerate(short):
        if i > 0 and ch.isupper() and not short[i - 1].isupper() and short[i - 1] != " ":
            out.append(" ")
        out.append(ch)
    return "".join(out)


def recipe_display_name(recipe_prefix: str, piece_name: str) -> str:
    # ("Chop", "FishPieces") -> "Chop Fish Pieces"
    return f"{recipe_prefix} {humanize_piece_name(piece_name)}"


def load_definitions_sheet(xlsx: Path) -> list[dict]:
    return load_xlsx_rows(xlsx, "definitions")


def load_evolved_matrix(xlsx: Path):
    """
    Reads the 'evolved' sheet and returns:
      - evolved: { recipe_name: [fullType, ...] }
      - keys: set of (module, piece_name)
      - per_item: { (module, piece_name): "Recipe:Value;..." }
    """
    rows = load_xlsx_rows(xlsx, "evolved")
    evolved: dict[str, list[str]] = {}
    keys: set[tuple[str, str]] = set()
    per_item: dict[tuple[str, str], str] = {}

    if not rows:
        return evolved, keys, per_item

    recipe_columns = [c for c in rows[0].keys() if c not in ("module", "piece_name")]

    for row in rows:
        key = (row["module"], row["piece_name"])
        keys.add(key)

        full_type = f"{row['module']}.{row['piece_name']}"
        parts = []

        for recipe in recipe_columns:
            val = (row.get(recipe) or "").strip()
            if val and val != "-":
                evolved.setdefault(recipe, []).append(full_type)
                parts.append(f"{recipe}:{val}")

        if parts:
            per_item[key] = ";".join(parts)

    return evolved, keys, per_item


def emit_item(row: dict, evolved_per_item: dict) -> str:
    lines: list[str] = []

    lines.append(f"    item {row['piece_name']}")
    lines.append("    {")
    lines.append(f"        DisplayName        = {display_name_key(row)},")
    lines.append(f"        Icon               = {row['icon']},")
    lines.append(f"        DisplayCategory    = {row['display_category']},")
    lines.append("        Type               = Food,")
    lines.append("")

    # Core consumption
    lines.append("        // Core consumption semantics (fungible pile)")
    lines.append(f"        HungerChange       = {row['hunger_change']},")
    lines.append(f"        BaseHunger         = {row['base_hunger']},")
    lines.append(f"        ThirstChange       = {row['thirst_change']},")
    lines.append("")

    # Nutrition (optional)
    if dash_is_none(row.get('calories', '-')) is not None:
        lines.append("        // Nutrition (base-hunger–scaled; conservative)")
        lines.append(f"        Calories           = {row['calories']},")
        lines.append(f"        Proteins           = {row['proteins']},")
        lines.append(f"        Lipids             = {row['lipids']},")
        lines.append(f"        Carbohydrates      = {row['carbohydrates']},")
        lines.append("")

    # Weight
    lines.append("        // Weight / encumbrance")
    lines.append(f"        Weight             = {row['weight']},")
    lines.append(f"        WeightFull         = {row['weight_full']},")
    lines.append(f"        WeightEmpty        = {row['weight_empty']},")
    lines.append("")

    # Freshness (optional)
    if dash_is_none(row.get('days_fresh', '-')) is not None:
        lines.append("        // Freshness / spoilage")
        lines.append(f"        DaysFresh          = {row['days_fresh']},")
        lines.append(f"        DaysTotallyRotten  = {row['days_totally_rotten']},")
        lines.append("")

    # Mood (optional)
    if dash_is_none(row.get('boredom_change', '-')) is not None:
        lines.append("        // Eating effects")
        lines.append(f"        BoredomChange      = {row['boredom_change']},")
        lines.append(f"        UnhappyChange      = {row['unhappy_change']},")
        lines.append("")

    # Metadata
    lines.append("        // Cooking / ingredient metadata")
    lines.append(f"        FoodType           = {row['food_type']},")
    lines.append(f"        Tags               = {row['tags']},")
    key = (row["module"], row["piece_name"])
    evolved_spec = evolved_per_item.get(key)
    if evolved_spec:
        lines.append(f"        EvolvedRecipe      = {evolved_spec},")
    if (row.get('dangerous_uncooked') or "").strip().lower() == "true":
        lines.append("        DangerousUncooked  = true,")
    lines.append("")
    lines.append("        // General flags")
    lines.append(f"        IsCookable         = {(row.get('is_cookable') or 'true').strip().lower()},")
    lines.append("        CanStoreWater      = false,")
    lines.append("    }")
    return "\n".join(lines)


def emit_chop_recipe(row: dict) -> str:
    if dash_is_none(row.get('chop_input_item_types', '-')) is None:
        return ""

    tools = (row.get('chop_required_tools') or "").replace("|", "/")

    lines: list[str] = []
    lines.append(f"    recipe Chop{row['piece_name']}")
    lines.append("    {")
    lines.append(f"        {row['chop_input_item_types']};1,")
    lines.append(f"        keep {tools},")
    lines.append(f"        Result      : {piece_full_type(row)}=1,")
    lines.append("        Time        : 50,")
    lines.append("        Category    : Cooking,")
    lines.append("        CanBeDoneFromFloor : true,")
    lines.append("        StopOnWalk    		: false,")
    lines.append("        NeedToBeLearn 		: false,")
    lines.append("        OnGiveXP    : Recipe.OnGiveXP.Cooking10,")
    lines.append("        OnCreate    : Recipe.OnCreate.KitchenConsolidation_Chop,")
    lines.append("    }")
    return "\n".join(lines)


def emit_combine_recipe(row: dict) -> str:
    lines: list[str] = []
    lines.append(f"    recipe Combine{row['piece_name']}")
    lines.append("    {")
    lines.append(f"        {piece_full_type(row)};2,")
    lines.append(f"        Result : {piece_full_type(row)}=1,")
    lines.append("        Time : 50,")
    lines.append("        Category : Cooking,")
    lines.append("        CanBeDoneFromFloor : true,")
    lines.append("        StopOnWalk    		: false,")
    lines.append("        NeedToBeLearn 		: false,")
    lines.append("        OnCanPerform : Recipe.OnCanPerform.KitchenConsolidation_Combine_OnCanPerform,")
    lines.append("        OnCreate : Recipe.OnCreate.KitchenConsolidation_Combine_OnCreate,")
    lines.append("    }")
    return "\n".join(lines)


def emit_evolvedrecipes(evolved: dict) -> str:
    lines: list[str] = []
    lines.append("// ------------------------------------------------------------------")
    lines.append("// Kitchen Consolidation – Evolved Recipe Extensions")
    lines.append("// AUTO-GENERATED FILE – DO NOT EDIT BY HAND")
    lines.append("// Source: food-evolved-vanilla.csv")
    lines.append("// ------------------------------------------------------------------")
    lines.append("")    
    lines.append("module Base {")

    for recipe_name in sorted(evolved.keys()):
        lines.append("    // =====================")
        lines.append(f"    // {recipe_name.upper()}")
        lines.append("    // =====================")
        lines.append(f"    evolvedrecipe {recipe_name}")
        lines.append("    {")
        for full_type in sorted(set(evolved[recipe_name])):
            lines.append(f"        Item {full_type},")
        lines.append("    }")
        lines.append("")
    lines.append("}")
    return "\n".join(lines)


def emit_item_translations(rows: list[dict], evolved_per_item: dict) -> str:
    lines = []
    lines.append("# --------------------------------------------------")
    lines.append("# Kitchen Consolidation – Item Names (EN)")
    lines.append("# AUTO-GENERATED FILE – DO NOT EDIT BY HAND")
    lines.append("# --------------------------------------------------")
    lines.append("")
    lines.append("ItemName_EN = {")

    for row in rows:
        base_name = humanize_piece_name(row["piece_name"])
        key = display_name_key(row)
        # IMPORTANT: emit as string key
        lines.append(f"    [\"{key}\"] = \"{base_name}\",")

    lines.append("}")
    return "\n".join(lines)


def emit_recipe_translations(rows: list[dict], containerized_rows: list[tuple[str, list[str]]]) -> str:
    lines = []
    seen = set()

    lines.append("# --------------------------------------------------")
    lines.append("# Kitchen Consolidation – Recipe Names (EN)")
    lines.append("# AUTO-GENERATED FILE – DO NOT EDIT BY HAND")
    lines.append("# --------------------------------------------------")
    lines.append("")
    lines.append("Recipes_EN = {")

    # Pieces recipes
    for row in rows:
        piece = row["piece_name"]

        if dash_is_none(row.get("chop_input_item_types", "-")) is not None:
            key = f"Recipe_Chop{piece}"
            if key not in seen:
                display = recipe_display_name('Chop', piece)
                display = re.sub(r"\s*\d+$", "", display)
                lines.append(f"    [\"{key}\"] = \"{display}\",")
                seen.add(key)

        key = f"Recipe_Combine{piece}"
        if key not in seen:
            display = recipe_display_name('Combine', piece)
            display = re.sub(r"\s*\d+$", "", display)
            lines.append(f"    [\"{key}\"] = \"{display}\",")
            seen.add(key)

    # Containerized combine recipes
    for item_id, _ in sorted(containerized_rows, key=lambda x: x[0]):
        short = item_id.split(".")[-1]
        key = f"Recipe_Combine{short}"
        if key not in seen:
            value = f"Combine {humanize_item_id(item_id)}"
            # Trim trailing " Open"
            if value.endswith(" Open"):
                value = value[:-5]
            # Trim trailing digits (e.g. "Crisps2" -> "Crisps")
            value = re.sub(r"\s*\d+$", "", value)
            lines.append(f"    [\"{key}\"] = \"{value}\",")
            seen.add(key)

    lines.append("}")
    return "\n".join(lines)



# --- Subcommand CLI implementation ---

def cmd_generate(xlsx_path: Path, mod_root_arg: Path):
    mod_root = mod_root_arg / "media"
    if not mod_root.exists():
        die(f"Resolved mod media directory does not exist: {mod_root}")

    scripts_folder = mod_root / "scripts"
    scripts_folder.mkdir(parents=True, exist_ok=True)

    translate_en = mod_root / "lua" / "shared" / "Translate" / "EN"
    translate_en.mkdir(parents=True, exist_ok=True)

    # Deterministic ordering
    containerized_rows = load_containerized_sheet(xlsx_path)
    rows = sorted(load_definitions_sheet(xlsx_path), key=lambda r: r["piece_name"])
    containerized_rows = sorted(containerized_rows, key=lambda x: x[0])
    evolved, evolved_keys, evolved_per_item = load_evolved_matrix(xlsx_path)

    # Validate 1:1 correspondence
    item_keys = {(r["module"], r["piece_name"]) for r in rows}
    if item_keys != evolved_keys:
        missing = item_keys - evolved_keys
        extra = evolved_keys - item_keys
        if missing:
            die(f"Items missing from evolved sheet: {sorted(missing)}")
        if extra:
            die(f"Extra items in evolved sheet: {sorted(extra)}")

    # Generate consolidated pieces.txt
    pieces_lines: list[str] = []
    pieces_lines.append("module KitchenConsolidation")
    pieces_lines.append("{")
    pieces_lines.append("    imports")
    pieces_lines.append("    {")
    pieces_lines.append("        Base,")
    pieces_lines.append("    }")
    pieces_lines.append("")
    for row in rows:
        pieces_lines.append(emit_item(row, evolved_per_item))
        pieces_lines.append("")
        chop = emit_chop_recipe(row)
        if chop:
            pieces_lines.append(chop)
            pieces_lines.append("")
        pieces_lines.append(emit_combine_recipe(row))
        pieces_lines.append("")
    pieces_lines.append("}")

    (scripts_folder / "pieces.txt").write_text("\n".join(pieces_lines), encoding="utf-8")

    # Generate containerized.txt
    cont_lines: list[str] = []
    cont_lines.append("module KitchenConsolidation")
    cont_lines.append("{")
    cont_lines.append("    imports")
    cont_lines.append("    {")
    cont_lines.append("        Base,")
    cont_lines.append("    }")
    cont_lines.append("")
    for item_id, _ in sorted(containerized_rows, key=lambda x: x[0].lower()):
        cont_lines.append(emit_containerized_combine_recipe(item_id))
        cont_lines.append("")
    cont_lines.append("}")

    (scripts_folder / "containerized.txt").write_text("\n".join(cont_lines), encoding="utf-8")

    # EN translations
    (translate_en / "ItemName_EN.txt").write_text(
        emit_item_translations(rows, evolved_per_item), encoding="utf-8"
    )
    (translate_en / "Recipes_EN.txt").write_text(
        emit_recipe_translations(rows, containerized_rows), encoding="utf-8"
    )

    # Generate RecipeContainerized.lua
    lua_path = mod_root / "lua" / "server" / "RecipeContainerized.lua"
    lua_path.parent.mkdir(parents=True, exist_ok=True)
    lua_path.write_text(
        emit_containerized_lua(containerized_rows), encoding="utf-8"
    )


def cmd_translate(mod_root_arg: Path, langs: List[str] = LANGS_DEFAULT):
    if OpenAI is None:
        die("openai SDK not installed. Run: pip install openai")

    mod_root = mod_root_arg / "media"
    base_dir = mod_root / "lua" / "shared" / "Translate"
    en_dir = base_dir / "EN"
    if not en_dir.exists():
        die(f"EN translation folder not found: {en_dir}")

    client = OpenAI()

    templates = []
    for domain, prefix in DOMAIN_FILES:
        src = en_dir / f"{prefix}_EN.txt"
        domain_name, table_name, entries = read_domain_table(src)
        templates.append((domain_name, entries))

    for lang in langs:
        if lang == "EN":
            continue
        out_dir = base_dir / lang
        for domain, entries in templates:
            out_path = out_dir / f"{domain}_{lang}.txt"
            prompt = build_translation_prompt(domain, lang, entries)
            translated = call_openai_translate(client, "gpt-5.2", prompt)
            write_domain_table(out_path, domain, lang, translated)
            print(f"Wrote: {out_path}")
            time.sleep(0.5)


def cmd_verify(xlsx_path: Path):
    errors: list[str] = []
    warns: list[str] = []

    # --- Load authoritative data ---
    rows = load_definitions_sheet(xlsx_path)
    evolved, evolved_keys, evolved_per_item = load_evolved_matrix(xlsx_path)
    containerized_rows = load_containerized_sheet(xlsx_path)

    # --- 2A: definitions <-> evolved must match exactly ---
    item_keys = {(r["module"], r["piece_name"]) for r in rows}

    if item_keys != evolved_keys:
        missing = item_keys - evolved_keys
        extra = evolved_keys - item_keys
        if missing:
            errors.append(f"Items missing from evolved sheet: {sorted(missing)}")
        if extra:
            errors.append(f"Extra items in evolved sheet: {sorted(extra)}")

    # --- 2D: containerized references must be valid ---
    defined_fulltypes = {f"{r['module']}.{r['piece_name']}" for r in rows}
    for item_id, _ in containerized_rows:
        if item_id.startswith("Base."):
            continue
        if item_id not in defined_fulltypes:
            errors.append(f"Containerized item not defined in definitions: {item_id}")

    # --- 2C: expected generated files + keys must exist ---
    # We know exactly what should be generated
    # Look in ./media relative to CWD
    mod_media = Path("media")

    scripts_dir = mod_media / "scripts"
    translate_en = mod_media / "lua" / "shared" / "Translate" / "EN"

    if not scripts_dir.exists():
        errors.append("Missing generated folder: media/scripts")
    else:
        pieces_file = scripts_dir / "pieces.txt"
        if not pieces_file.exists():
            errors.append("Missing generated file: media/scripts/pieces.txt")

        containerized_file = scripts_dir / "containerized.txt"
        if containerized_rows and not containerized_file.exists():
            errors.append("Missing generated file: media/scripts/containerized.txt")

    # --- Translation completeness (EN is authoritative) ---
    itemname_en = translate_en / "ItemName_EN.txt"
    recipes_en = translate_en / "Recipes_EN.txt"

    if not itemname_en.exists():
        errors.append("Missing translation file: ItemName_EN.txt")
    if not recipes_en.exists():
        errors.append("Missing translation file: Recipes_EN.txt")

    # Parse EN translation tables if present
    en_item_keys: set[str] = set()
    en_recipe_keys: set[str] = set()

    if itemname_en.exists():
        _, _, entries = read_domain_table(itemname_en)
        en_item_keys = set(entries.keys())

    if recipes_en.exists():
        _, _, entries = read_domain_table(recipes_en)
        en_recipe_keys = set(entries.keys())

    # Expected item keys
    for r in rows:
        key = display_name_key(r)
        if key not in en_item_keys:
            errors.append(f"Missing EN item translation key: {key}")

    # Expected recipe keys
    for r in rows:
        piece = r["piece_name"]
        if dash_is_none(r.get("chop_input_item_types", "-")) is not None:
            k = f"Recipe_Chop{piece}"
            if k not in en_recipe_keys:
                errors.append(f"Missing EN recipe translation key: {k}")
        k = f"Recipe_Combine{piece}"
        if k not in en_recipe_keys:
            errors.append(f"Missing EN recipe translation key: {k}")

    for item_id, _ in containerized_rows:
        short = item_id.split(".")[-1]
        k = f"Recipe_Combine{short}"
        if k not in en_recipe_keys:
            errors.append(f"Missing EN recipe translation key: {k}")

    # --- 4: Non-EN translations are warnings only ---
    base_translate = mod_media / "lua" / "shared" / "Translate"
    for lang in LANGS_DEFAULT:
        if lang == "EN":
            continue
        lang_dir = base_translate / lang
        if not lang_dir.exists():
            warns.append(f"Missing language directory: {lang}")
            continue

        for fname in ("ItemName_" + lang + ".txt", "Recipes_" + lang + ".txt"):
            path = lang_dir / fname
            if not path.exists():
                warns.append(f"Missing translation file: {lang}/{fname}")
                continue

            _, _, entries = read_domain_table(path)
            # Warn on missing keys only
            if fname.startswith("ItemName"):
                for k in en_item_keys:
                    if k not in entries:
                        warns.append(f"{lang} missing item translation: {k}")
            else:
                for k in en_recipe_keys:
                    if k not in entries:
                        warns.append(f"{lang} missing recipe translation: {k}")

    # --- Emit results ---
    for e in errors:
        print(f"VERIFY ERROR: {e}", file=sys.stderr)
    for w in warns:
        print(f"VERIFY WARN : {w}", file=sys.stderr)

    if errors:
        sys.exit(2)

    print("VERIFY OK")


def main():
    parser = argparse.ArgumentParser(prog="gen-items.py")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_gen = sub.add_parser("generate", help="Generate scripts and EN translations")
    p_gen.add_argument("xlsx", type=Path)
    p_gen.add_argument("mod_root", type=Path)

    p_tr = sub.add_parser("translate", help="Generate non-EN translations")
    p_tr.add_argument("mod_root", type=Path)

    p_v = sub.add_parser("verify", help="Validate data without writing files")
    p_v.add_argument("xlsx", type=Path)

    args = parser.parse_args()

    if args.cmd == "generate":
        cmd_generate(args.xlsx, args.mod_root)
    elif args.cmd == "translate":
        cmd_translate(args.mod_root)
    elif args.cmd == "verify":
        cmd_verify(args.xlsx)


if __name__ == "__main__":
    main()
